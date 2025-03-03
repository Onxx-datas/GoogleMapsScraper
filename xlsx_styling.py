import openpyxl # Library to handle (.xlsx) Excel file
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # importing styling modules from Openpyxl

def apply_styles(file_name): # Creating funciton to call it on main script that named "apply_styles"
    wb = openpyxl.load_workbook(file_name) # Loading Excel workbook from actual file
    ws = wb.active # Using workbook "wb" as "ws"
    header_font = Font(bold=True, color="FFFFFF") # Making header texts Bold and White
    header_fill = PatternFill(start_color="0073CF", end_color="0073CF", fill_type="solid") # Making background color as  Blue
    border_style = Border( # Applying styles to each 4 borders...
        left=Side(style="thin"), # Making left side of border thin
        right=Side(style="thin"), # Making right side of border thin
        top=Side(style="thin"), # Making top side of border thin
        bottom=Side(style="thin")) # Making tpo side of border thin
    for col in range(1, ws.max_column + 1): # Increasing number of rows
        cell = ws.cell(row=1, column=col)
        cell.font = header_font # Font of the cell will be as a headers color
        cell.fill = header_fill # And background color will be as a header fill
        cell.alignment = Alignment(horizontal="center", vertical="center") # Making text's alignment horizontally and vertically centered 
        cell.border = border_style # Applying border style into each cell
    column_widths = [71, 29, 13, 8, 11] # Changing each column's width (Column by column)
    for i, width in enumerate(column_widths, start=1): # Starting column widht form 1 column
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width # Set the width of column 'i' to the specified value
    for row in ws.iter_rows(): # Iterate over all rows in the worksheet
        for cell in row: # Iterate over each cell in the current row
            cell.border = border_style # Apply the specified border style to the cell
    wb.save(file_name) # Saving all changes into file name parameter
